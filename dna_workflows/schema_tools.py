import sdtables
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from dna_workflows import package_tools

# manifest = 'manifest.yml'
# xl_out = 'example.xlsx'


def create_new_workbook():
    _wb = Workbook()
    _ws = _wb.active
    _wb.remove(_ws)

    return _wb


def load_xl_wf_db(excel_db, flatten=False, fill_empty=False):
    _data = sdtables.load_xl_db(excel_db, flatten=flatten)

    return _data


def build_module_schema(_wb, _modules, user_data=None):
    """ Builds Excel tables from schema based on module manifest.

    :param _wb: (object) An openpyxl workbook object
    :param modules: (dict) Dictionary describing the project modules loaded from schema.yml
    :param user_data: (dict) In the case that there is existing data to preserve we can provide the existing table data

    :returns: an updated openpyxl workbook object """

    for module in _modules:
        module_doc = package_tools.get_module_doc(module)

        if 'schemas' in module_doc['module']:
            ws = _wb.create_sheet(module_doc['module']['name'])
            ws.sheet_properties.tabColor = "009900"
        else:
            continue

        if 'description' in module_doc['module'].keys():
            ws.append({2: module_doc['module']['description']})
            ws['B1'].fill = PatternFill("solid", fgColor="b6e5eb")
            ws['B1'].alignment = Alignment(wrapText=True, vertical='top')
            ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=11)

        for module_schema in module_doc['module']['schemas'].keys():
            module_name = module_doc['module']['name']
            name = '{}.{}.{}'.format(module_schema, 'schema', module_name)
            schema_doc = module_doc['module']['schemas'][module_schema]
            if user_data is not None:
                if name in user_data.keys():
                    data = user_data[name]
                    sdtables.add_schema_table_to_worksheet(ws, name, schema_doc, data=data, table_style='TableStyleMedium2')
                else:
                    sdtables.add_schema_table_to_worksheet(ws, name, schema_doc, table_style='TableStyleMedium2')
            elif 'data' in module_doc['module']:
                if module_schema in module_doc['module']['data'].keys():
                    data = module_doc['module']['data'][module_schema]
                    sdtables.add_schema_table_to_worksheet(ws, name, schema_doc, data=data, table_style='TableStyleMedium2')
                else:
                    sdtables.add_schema_table_to_worksheet(ws, name, schema_doc, table_style='TableStyleMedium2')
            else:
                sdtables.add_schema_table_to_worksheet(ws, name, schema_doc, table_style='TableStyleMedium2')

    return _wb


def build_workflow_task_sheet(_wb, _modules, user_data=None):
    """ Builds the workflows cover sheet with a table containing available tasks
    based on the project manifest loaded from schema.yml.

    :param _wb: (object) An openpyxl workbook object
    :param _modules: (dict) Dictionary describing the project modules loaded from schema.yml
    :param user_data: (dict) In the case that there is existing data to preserve we can provide the existing table data

    :returns: an updated openpyxl workbook object """

    _manifest = package_tools.load_install_manifest()
    from dna_workflows import wf_engine
    wf_doc = wf_engine.get_module_definition()
    _ws = _wb.create_sheet("workflows", 0)
    _ws.sheet_properties.tabColor = "0080FF"

    for _package_name, _package_meta in _manifest['manifest'].items():
        _module_list = []
        if _package_meta['type'] == 'module':
            _module_list.append(_package_name)
            _schema_suffix = _package_name
        elif _package_meta['type'] == 'bundle':
            _schema_suffix = 'bundle.{}'.format(_package_name)
            for _bundle_module in _package_meta['provides']:
                _module = '{}.{}'.format(_package_name, _bundle_module)
                _module_list.append(_module)

        # Build a list of modules and tasks based on the manifest
        methods = []
        for _module in _module_list:
            module_doc = package_tools.get_module_doc(_module)

            for m in module_doc['module']['methods']:
                methods.append(m)

        _table_name = 'workflow.{}'.format(_schema_suffix)
        if user_data is not None:
            user_methods = get_data_from_schema(_table_name, user_data)
            for _m in methods:
                for row in user_methods:
                    if _m['module'] == row['module'] and _m['task'] == row['task']:
                        _m['status'] = row['status']

        wf_schema = wf_doc['module']['schemas']['workflow']
        sdtables.add_schema_table_to_worksheet(_ws, _table_name, wf_schema, data=methods, table_style='TableStyleLight14')

        # Add conditional formatting to workflow worksheet
        for table in _ws.tables.values():
            if _table_name == table.name:
                _tdef = table.ref
                red_fill = PatternFill(bgColor="9da19e")
                dxf = DifferentialStyle(fill=red_fill)
                r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
                _formula = '${}="disabled"'.format(_tdef.split(':')[0])
                r.formula = [_formula]
                _ws.conditional_formatting.add(_tdef, r)

    return _wb


def get_data_from_schema(_schema_name, _db):
    if _schema_name in _db.keys():
        return _db[_schema_name]
    else:
        return []


def validate_module_schema(_workflow_db, stdout=False):
    """ Validates the pre-loaded workflow DB against the module schema.

    :param stdout: Print results to STDOUT
    :param _workflow_db: (object) A DNA Workflows db pre-loaded from either Excel or YAML

    :returns: Boolean depending on validation results """
    _workflow_db['workflow'] = package_tools.transform_wf_tasks(_workflow_db)
    _modules = []
    for _task in _workflow_db['workflow']:
        if 'enabled' in _task['status'] and _task['module'] not in _modules:
            _modules.append(_task['module'])

    _results = []
    for _module in _modules:
        module_doc = package_tools.get_module_doc(_module)
        module_name = module_doc['module']['name']
        print('VALIDATING module: {}'.format(module_name))
        if 'schemas' in module_doc['module']:
            for _schema, _properties in module_doc['module']['schemas'].items():
                _dnawf_schema_name = '{}.schema.{}'.format(_schema, module_name)
                if _dnawf_schema_name in _workflow_db.keys():
                    print('VALIDATING schema: {}'.format(_dnawf_schema_name))
                    _result = sdtables.validate_data(_properties, _workflow_db[_dnawf_schema_name])
                    if stdout:
                        print_validation_results(_result)
                    else:
                        _results.append(_result)
                else:
                    print('scheam {} not found'.format(_dnawf_schema_name))

    return _results


def print_validation_results(_results):
    print('RESULT: {}'.format(_results['result']))
    for _r in _results['details']:
        if _r['result'] != 'OK':
            print('Row {} contains validation errors:'.format(_r['row']+1))
            print('------------------ERROR-----------------'.center(50))
            print(_r['result'])
            print('--------------------------------------'.center(50))

# if __name__ == "__main__":
#     schema = yaml.load(open(manifest, 'r'), Loader=yaml.SafeLoader)
#     wb = create_new_workbook()
#     wb = build_module_schema(wb, schema)
#     wb = build_workflow_task_sheet(wb, schema)
#     wb.save(xl_out)
