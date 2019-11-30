from jinja2 import Template
import argparse
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import os


def main():
    module = args.add_workflow

    # Create module directory
    try:
        os.mkdir(module)
    except OSError as e:
        print('Creation of the directory {} failed: {}'.format(module, e))
    else:
        print('Successfully created the directory {}'.format(module))

    # Write module template files
    tm = Template(workflow_template)
    workflow = tm.render(module=module)
    write_to_file(module, 'workflow.py', workflow)

    tm = Template(init_file_template)
    init_file = tm.render()
    write_to_file(module, '__init__.py', init_file)

    # Add module to workflows DB
    add_module_to_excel(workflow_db, module)


def add_module_to_excel(workflow_db, module):
    wb = openpyxl.load_workbook(filename=workflow_db)
    if module in wb.sheetnames:
        print('module {} already exists. If this is an error:'.format(module))
        print('1. Delete the worksheet named {}'.format(module))
        print('2. Remove the workflow entry from the workflows sheet for {}'.format(module))
        print('3. Remove the directory ./{} and all of its contents'.format(module))

    else:
        ws = wb.create_sheet(module)
        ws.append(['Stage', 'Status', 'Function', 'Documentation'])
        ws.append([1, 'enabled', 'hello_world', 'Prints Hello World!!'])
        ws.move_range("A1:D2", rows=1, cols=1)
        table_name = 'control?{}?Function'.format(module)
        tab = Table(displayName=table_name, ref="B2:E3")
        style = TableStyleInfo(name="TableStyleMedium3", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        _workflow_sheet = wb['workflows']
        for table in _workflow_sheet._tables:
            if table.name == 'workflow':
                print(get_table_length(table))
                _workflow_id = get_table_length(table) + 1
                row = [_workflow_id, 'enabled', module, 'Update this workflow documentation']
                add_row_to_table(_workflow_sheet, table, row)

        wb.save(workflow_db)


def add_row_to_table(worksheet, table, row):
    _table_start, _table_end = table.ref.split(':')
    _table_start = get_cell_coordinates(_table_start)
    _table_end = get_cell_coordinates(_table_end)
    _new_table_end = (_table_end[0], _table_end[1] + 1)
    _new_table_end = get_cell_ref_from_coordinate(_new_table_end)
    _new_ref = '{}:{}'.format(table.ref.split(':')[0], _new_table_end)
    table.ref = _new_ref

    for idx, val in enumerate(row):
        worksheet.cell(row=_table_end[1] + 1, column=_table_start[0] + idx).value = val


def get_cell_coordinates(ref):
    xy = openpyxl.utils.cell.coordinate_from_string(ref)
    col = openpyxl.utils.cell.column_index_from_string(xy[0])
    row = xy[1]

    return col, row


def get_cell_ref_from_coordinate(coordinate):
    x = coordinate[0]
    openpyxl.utils.cell.get_column_letter(x)
    ref = '{}{}'.format(openpyxl.utils.cell.get_column_letter(x), coordinate[1])

    return ref


def get_table_length(table):
    _table_start, _table_end = table.ref.split(':')
    _table_start = get_cell_coordinates(_table_start)
    _table_end = get_cell_coordinates(_table_end)

    return _table_end[1] - _table_start[1]


def write_to_file(path, file, data):
    _path_to_file = '{}/{}'.format(path, file)
    f = open(_path_to_file, "w")
    f.write(data)
    f.close()


# Check for args
workflow_db = 'dna_workflow_db.xlsx'

parser = argparse.ArgumentParser()
parser.add_argument("--add-workflow", help="Adds a new workflow template")
parser.add_argument("--delete-workflow", help="Deletes a workflow from the DB")
args = parser.parse_args()

workflow_template = '''
import logging
import common

logger = logging.getLogger('main.{{ module }}')


def hello_world(api, workflow_dict):
    logger.info('reports::hello_world')
    for key, value in workflow_dict.items():
        logger.info('Found table: {} with rows: '.format(key))
        for row in value:
            logger.info('{}'.format(row))

'''

init_file_template = '''
"""
This is a python init file generated from template.  Replace this text with your workflow description
"""

__version__ = "0.0.1"

from .workflow import *
'''

if __name__ == "__main__":
    main()